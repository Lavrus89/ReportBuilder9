import xlrd, xlwt
import os.path, datetime
from openpyxl.utils.cell import get_column_letter
import tkinter as tk
from tkinter.filedialog import askopenfilename, askopenfilenames, asksaveasfilename
# import pprint
# pp = pprint.PrettyPrinter(indent=4)

# files = ('год 27.09.xls', 'месяц 2709.xls')
# report = 'Отчет 2709.xls'


class Report:
    def __init__(self, report, files):
        self.report = report
        self.files = files
        self.goroda = self._goroda() #словарь клиент:город из отчета за большой период
        self.goods_props = self._goods_props() # список всех свойств товаров в отчетах
        self.clients = self._clients() # Список клиентов из шахматок
        self.clear_report = self._clear_report() #пустой отчет со всеми клиентами и свойствами
        self._fill_report()
        self.itog = self._itog()

    def _goroda(self):
        ws = xlrd.open_workbook(self.report, encoding_override='1251').sheet_by_index(0)
        return {ws.cell(row, 0).value: ws.cell(row+1, 0).value for row in range(9, ws.nrows-1)}

    def _goods_props(self):
        props = {}
        for file in self.files:
            ws = xlrd.open_workbook(file, encoding_override='1251').sheet_by_index(0)
            data = [ws.cell(5, col).value for col in range(1, ws.ncols-2)]
            for i in data:
                props[i] = '0'
        return tuple(props.keys())

    def _clients(self):
        clients = {}
        for file in self.files:
            ws = xlrd.open_workbook(file, encoding_override='1251').sheet_by_index(0)
            data = [ws.cell(row, 0).value for row in range(6, ws.nrows-2)]
            for i in data:
                clients[i] = '0'
        return tuple(clients.keys())

    def _clear_report(self):
        report = {client:{file:{prop:' ' for prop in self.goods_props} for file in self.files} for client in self.clients}
        for key in report.keys():
            try:
                report[key]['gorod'] = self.goroda[key]
            except KeyError:
                report[key]['gorod'] = ''
        return report

    def _fill_report(self):
        for file in self.files:
            period = self.get_period(file)
            ws = xlrd.open_workbook(file, encoding_override='1251').sheet_by_index(0)
            for row in range(6, ws.nrows-2):
                for col in range(1, ws.ncols-2):
                    self.clear_report[ws.cell(row, 0).value][file][ws.cell(5, col).value] = ws.cell(row, col).value / period if ws.cell(row, col).value != ' ' else ' '

    def get_report(self):
        return self.clear_report

    def get_period(self, file):
        date = xlrd.open_workbook(file, encoding_override="cp1251").sheet_by_index(0).cell(3, 0).value
        date_1 = datetime.datetime.strptime(date[2:10], '%d.%m.%y')
        date_2 = datetime.datetime.strptime(date[14:], '%d.%m.%y')
        return (date_2 - date_1).days / 30

    def _itog(self):
        itog = {file:{prop: 0 for prop in self.goods_props} for file in self.files}
        for file in self.files:
            for client in self.clear_report.keys():
                for prop in self.goods_props:
                    itog[file][prop] += self.clear_report[client][file][prop] if self.clear_report[client][file][prop] != ' ' else 0
        # pp.pprint(itog)
        return itog

    def save_report(self, filename):
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Отчёт')
        ws.col(0).width = 5000
        ws.row(0).height = 2000
        ws.col(1).width = 5000
        ws.col(2).width = 12700
        # Форматы
        header_format = xlwt.Style.easyxf(
            "borders: left thin, right thin, top thin, bottom thin; font: bold True; align: horiz center; alignment: wrap True")
        text_format = xlwt.Style.easyxf("borders: left thin, right thin, top thin, bottom thin")
        digit_format = xlwt.Style.easyxf("borders: left thin, right thin, top thin, bottom thin")
        digit_format.num_format_str = "# ##0"

        digit_format2 = xlwt.Style.easyxf("borders: left thin, right thin, top thin, bottom thin")
        digit_format2.num_format_str = "# ##0.00"
        # Шапка
        ws.write(0, 0, 'Город', header_format)
        ws.write(0, 1, 'Файл', header_format)
        ws.write(0, 2, 'Покупатель/Св-во ТМЦ', header_format)
        for i in range(len(self.goods_props)):
            ws.write(0, 4+i, self.goods_props[i], header_format)
        # Итоги
        for i in range(len(self.files)):
            ws.write(1+i, 1, os.path.basename(self.files[i])[:-4], text_format)
            ws.write(1+i, 2, self.get_period(self.files[i]), text_format)
            ws.write(1+i, 3, xlwt.Formula(f"SUM(E{i + 2}:HQ{i + 2})"), digit_format)
            for k in range(len(self.goods_props)):
                # ws.write(1+i, 4+k, 1)
                # column = get_column_letter(5+k)
                # ws.write(1+i, 4+k, xlwt.Formula(f'SUMIF(B{2+len(self.files)}:B30000;B{2+i};{column}{2+len(self.files)}:{column}30000)'), digit_format)
                ws.write(1 + i, 4 + k, self.itog[self.files[i]][self.goods_props[k]], digit_format)
        # основные данные
        row = 1 + len(self.files)
        for client in self.clear_report.keys():
            for file in self.files:
                ws.write(row, 0, self.clear_report[client]['gorod'], text_format)
                ws.write(row, 1, os.path.basename(file)[0:-4], text_format)
                ws.write(row, 2, client, text_format)
                ws.write(row, 3, xlwt.Formula(f"SUM(E{row + 1}:HQ{row + 1})"), digit_format)
                for i in range(len(self.goods_props)):
                    ws.write(row, 4 + i, self.clear_report[client][file][self.goods_props[i]], digit_format)
                row += 1

        wb.save(filename)

def select_sh():
    files = askopenfilenames()
    files2.set('')
    for file in files:
        files2.set(files2.get() + '\n' + file)

def select_sv():
    file = askopenfilename()
    file1.set(file)

def save_report():
    save_filename = asksaveasfilename(filetypes=[("xls file", ".xls")], defaultextension=".xls")
    files = files2.get().split('\n')[1:]
    report = file1.get()
    my_rep = Report(report, files)
    my_rep.save_report(save_filename)

root = tk.Tk()
root.geometry("600x400")
root.title("Шахматки")
root.resizable(False, False)


lsb2_text = """Выберите шахматки:
(Операции -> Отчеты -> Торговая шахматка -> 
по горизонтали - свойства товаров
по вертикали покупатели
в множественном фильтре выбираем клиентов 
по Свойства покупателей:Менеджер по продажам
-> выбираем период  -> сформировать  -> сохранить в xls
-> и так по всем периодам какие нужны )"""
lab2 = tk.Label(root, text=lsb2_text)
lab2.pack()
but2 = tk.Button(root, text="Обзор...", command=select_sh)
but2.pack()
sh_files = []
files2 = tk.StringVar()
txt2 = tk.Label(root, textvariable=files2)
txt2.pack()

lab1_text = """Выберите отчет со свойствами клиентов:
(Очеты -> Отчет по продажам ТМЦ со свойствами -> выбираем период
с начала самой ранней шахматки по текщую дату, 
в детализации: Покупатель, Свойство покупателя -> Сформировать
 -> Сохраняем в xls)"""
lab1 = tk.Label(root, text=lab1_text)
lab1.pack()
but1 = tk.Button(root, text="Обзор...", command=select_sv)
but1.pack()
file1 = tk.StringVar()
txt1 = tk.Label(root, textvariable=file1)
txt1.pack()


but3 = tk.Button(root, text="Сохранить отчет как...", command=save_report)
but3.pack()

root.mainloop()
